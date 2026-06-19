from flask_cors import CORS
from dotenv import load_dotenv
from flask import Flask, jsonify, request, send_file
import boto3
import pandas as pd
import io
import os
from pathlib import Path
from boto3.dynamodb.conditions import Attr
import calendar

API_DIR = Path(__file__).resolve().parent
load_dotenv(API_DIR / ".env")
load_dotenv(API_DIR.parent / ".env", override=True)
# Set REGION from environment variable or default to "us-east-1"
REGION = os.getenv("REGION", "us-east-1")
# --- Auth-related imports ---
from werkzeug.security import generate_password_hash, check_password_hash
import uuid
import psycopg2
from psycopg2.extras import RealDictCursor
import secrets
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo          # built-in on Python 3.9+
from botocore.exceptions import ClientError
USERS = {}

RESET_TOKENS = {}  # email -> { token, expires_at }

app = Flask(__name__)
CORS(app, supports_credentials=True, origins=os.getenv("FRONTEND_URL", "*"))

import smtplib
from email.mime.text import MIMEText

DATE_FORMATS = (
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%YT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
)

STATUS_MAP = {
    -1: "Booked",
    -2: "Max Reached Out",
    -4: "No Selection Made",
    -3: "No Available Slots",
    -5: "Failed",
     1: "Reached Once",
     2: "Reached Twice",
}

DEPARTMENT_MAP = {
    "6": "OHCC Adult",
    "18": "OHCC HAART",
    "22": "OHCC Denham",
    "27": "OHCC Baranco",
    "28": "OHCC Colonial",
    "31": "OHCC Airline North",
    "33": "OHCC Baranco Televisit",
    "34": "OHCC Eunice",
    "35": "OHCC Capital Middle",
    "37": "OHCC Airline North Televisit",
    "38": "OHCC Sharon Hills",
    "39": "OHCC Claiborne",
    "40": "OHCC Jefferson Terrace",
    "41": "OHCC Progress Elementary",
    "42": "OHCC Jefferson Terrace Televisit",
    "43": "OHCC Claiborne Televisit",
    "44": "OHCC Eunice Televisit",
    "45": "OHCC Capital Middle Televisit",
    "46": "OHCC Sharon Hills Televisit",
    "47": "OHCC Progress Elementary Televisit"
}


def parse_datetime(value):
    if value is None:
        return None

    value = str(value).strip()
    if not value:
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    parsed = pd.to_datetime(value, errors='coerce')
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def parse_request_datetime(value):
    parsed = parse_datetime(value)
    if not parsed:
        raise ValueError("Date must be in YYYY-MM-DD HH:MM:SS format")
    return parsed


def scan_all_items(table):
    items = []
    response = table.scan()
    items.extend(response.get('Items', []))
    while 'LastEvaluatedKey' in response:
        response = table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
        items.extend(response.get('Items', []))
    return items


def scan_items_by_created_at_prefix(table, start, end):
    ranges = [
        (start.strftime("%m/%d/%Y %H:%M:%S"), end.strftime("%m/%d/%Y %H:%M:%S")),
        (start.strftime("%m/%d/%YT%H:%M:%S"), end.strftime("%m/%d/%YT%H:%M:%S")),
        (start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")),
        (start.strftime("%Y-%m-%dT%H:%M:%S"), end.strftime("%Y-%m-%dT%H:%M:%S")),
    ]

    expression = None
    for start_value, end_value in ranges:
        condition = Attr("created_at").between(start_value, end_value)
        expression = condition if expression is None else expression | condition

    items = []
    response = table.scan(FilterExpression=expression)
    items.extend(response.get('Items', []))
    while 'LastEvaluatedKey' in response:
        response = table.scan(
            FilterExpression=expression,
            ExclusiveStartKey=response['LastEvaluatedKey'],
        )
        items.extend(response.get('Items', []))
    return items


def filter_items_by_created_at(items, start, end):
    filtered = []
    for item in items:
        created_at = parse_datetime(item.get("created_at"))
        if created_at and start <= created_at <= end:
            filtered.append(item)
    return filtered


def scan_table_by_created_at(table, start, end):
    items = scan_items_by_created_at_prefix(table, start, end)
    filtered = filter_items_by_created_at(items, start, end)
    print(f"[📦] Retrieved {len(filtered)} of {len(items)} items from {table.name}")
    return filtered


def build_live_detail_records(start_date, end_date):
    dynamodb = boto3.resource('dynamodb', region_name=REGION)
    patient_table = dynamodb.Table(os.getenv("PATIENT_TABLE_NAME"))
    appointment_table = dynamodb.Table(os.getenv("APPOINTMENT_TABLE_NAME"))

    patients = scan_table_by_created_at(patient_table, start_date, end_date)
    appointments = scan_table_by_created_at(appointment_table, start_date, end_date)

    print(f"[✅] Patient records: {len(patients)} | Appointment records: {len(appointments)}")

    df_patients = pd.DataFrame(patients)
    df_appointments = pd.DataFrame(appointments)
    if df_patients.empty or df_appointments.empty:
        return []

    df_patients['old_appointment_id'] = df_patients['old_appointment_id'].astype(str)
    df_appointments['id'] = df_appointments['id'].astype(str)

    merged_df = pd.merge(df_appointments, df_patients, left_on='id', right_on='old_appointment_id', how='left')
    merged_df = merged_df.fillna('')

    merged_df['Status'] = merged_df['reach_out_count'].map(STATUS_MAP).fillna("Other")

    merged_df['created_at'] = pd.to_datetime(merged_df['created_at_x'], errors='coerce')
    merged_df['Date'] = merged_df['created_at'].dt.strftime('%Y-%m-%d')
    merged_df['Time'] = merged_df['created_at'].dt.strftime('%H:%M:%S')
    merged_df['Provider Name'] = merged_df['doctor_name']
    merged_df['Provider Name'] = merged_df['Provider Name'].fillna('').astype(str).str.strip().replace('', 'Unknown')

    merged_df['Patient ID'] = merged_df['patient_id_x']
    merged_df['Department'] = merged_df['department'].astype(str).map(DEPARTMENT_MAP).fillna('Unknown')

    merged_df = merged_df.sort_values(by='created_at', ascending=False)
    merged_df = merged_df.drop_duplicates(subset=['old_appointment_id'], keep='first')
    merged_df["Duplicate"] = ""

    columns = ['Patient ID', 'Date', 'Time', 'Department', 'Provider Name', 'Status', 'Duplicate', 'pt_language', 'reach_out_medium']
    return merged_df[columns].to_dict(orient='records')


def calculate_dashboard_metrics(records):
    unique_keys = set()
    booked_count = 0
    revenue_sum = 0

    for row in records:
        if row.get('Status') != 'Booked':
            continue

        key = f"{row.get('Patient ID')}_{row.get('Date')}_{row.get('Time')}_{row.get('Status')}"
        if key in unique_keys:
            continue

        unique_keys.add(key)
        booked_count += 1
        revenue_sum += 175 if (row.get('Date') or '') >= '2025-07-01' else 154

    total_sent = len(records)
    return {
        "total_sent": total_sent,
        "total_booked": booked_count,
        "percent_booked": (booked_count / total_sent) if total_sent else 0,
        "estimated_revenue": revenue_sum,
    }


def month_start_end(month_str):
    year, month = map(int, month_str.split("-"))
    start = datetime(year, month, 1, 0, 0, 0)
    end_day = calendar.monthrange(year, month)[1]
    end = datetime(year, month, end_day, 23, 59, 59)
    return start, end


def iter_months(start_month, end_month):
    start_year, start_mon = map(int, start_month.split("-"))
    end_year, end_mon = map(int, end_month.split("-"))
    current_year, current_mon = start_year, start_mon

    while (current_year, current_mon) <= (end_year, end_mon):
        yield f"{current_year}-{current_mon:02d}"
        current_mon += 1
        if current_mon == 13:
            current_year += 1
            current_mon = 1


def last_completed_month():
    clinic_tz = ZoneInfo(os.getenv("CLINIC_TIMEZONE", "America/Chicago"))
    today = datetime.now(clinic_tz).date()
    first_of_month = today.replace(day=1)
    previous_month_end = first_of_month - timedelta(days=1)
    return previous_month_end.strftime("%Y-%m")


def scan_ai_booking_items():
    dynamodb = boto3.resource('dynamodb', region_name=REGION)
    table = dynamodb.Table(os.getenv("AI_BOOKINGS_TABLE_NAME"))
    all_items = []
    resp = table.scan()
    all_items.extend(resp.get('Items', []))
    while "LastEvaluatedKey" in resp:
        resp = table.scan(ExclusiveStartKey=resp["LastEvaluatedKey"])
        all_items.extend(resp.get('Items', []))
    return all_items


def calculate_ai_show_rate(months, all_items=None):
    if all_items is None:
        all_items = scan_ai_booking_items()

    date_ranges = []
    for month_str in months:
        date_ranges.append(month_start_end(month_str))

    booked = 0
    kept = 0

    clinic_tz = ZoneInfo(os.getenv("CLINIC_TIMEZONE", "America/Chicago"))
    today = datetime.now(clinic_tz).date()

    for item in all_items:
        appt_date = parse_datetime(item.get("appointment_date"))
        if not appt_date:
            continue

        if appt_date.date() >= today:
            continue

        if not any(start <= appt_date <= end for (start, end) in date_ranges):
            continue

        booked += 1
        if (item.get("status") or "").strip() == "Kept":
            kept += 1

    show_rate = round((kept / booked) * 100, 1) if booked else 0.0
    return {
        "months": months,
        "booked": booked,
        "kept": kept,
        "show_rate": show_rate,
    }

@app.route('/request-reset', methods=['POST'])
def request_reset():
    data = request.json
    email = data.get('email')
    if not email:
        return jsonify({"error": "Email is required"}), 400

    token = secrets.token_urlsafe(16)
    expires_at = datetime.utcnow() + timedelta(minutes=15)
    RESET_TOKENS[email] = {"token": token, "expires_at": expires_at}

    # Compose email
    reset_link = f"https://yourdomain.com/reset-password?token={token}&email={email}"
    subject = "Your Password Reset Link"
    body = f"To reset your password, click the link below:\n\n{reset_link}\n\nThis link will expire in 15 minutes."
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = os.getenv("SMTP_EMAIL")
    msg["To"] = email

    try:
        with smtplib.SMTP(os.getenv("SMTP_SERVER"), int(os.getenv("SMTP_PORT"))) as server:
            server.starttls()
            server.login(os.getenv("SMTP_EMAIL"), os.getenv("SMTP_PASSWORD"))
            server.send_message(msg)
        return jsonify({"message": "Password reset email sent."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/reset-password', methods=['POST'])
def reset_password():
    data = request.json
    email = data.get('email')
    token = data.get('token')
    new_password = data.get('new_password')

    if not all([email, token, new_password]):
        return jsonify({"error": "Email, token, and new password are required"}), 400

    record = RESET_TOKENS.get(email)
    if not record or record['token'] != token or record['expires_at'] < datetime.utcnow():
        return jsonify({"error": "Invalid or expired token"}), 400

    password_hash = generate_password_hash(new_password)
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE users SET password_hash = %s WHERE email = %s", (password_hash, email))
        conn.commit()
        cur.close()
        conn.close()
        del RESET_TOKENS[email]
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"message": "Password has been reset successfully."})

@app.route('/live-details')
def live_details():
    input_start = request.args.get("start")
    input_end = request.args.get("end")

    print(f"[🕵️] live-details called with start={input_start} and end={input_end}")

    try:
        start_date = parse_request_datetime(input_start)
        end_date = parse_request_datetime(input_end)
    except Exception as e:
        return jsonify({"error": f"Invalid date format: {e}"}), 400

    if not start_date or not end_date:
        return jsonify({"error": "start and end dates are required"}), 400

    records = build_live_detail_records(start_date, end_date)
    print(f"[📊] Returning {len(records)} deduplicated results")
    return jsonify(records)

@app.route('/booked-by-provider')
def booked_by_provider():
    dynamodb = boto3.resource('dynamodb', region_name=REGION)
    patient_table_name = os.getenv("PATIENT_TABLE_NAME")
    appointment_table_name = os.getenv("APPOINTMENT_TABLE_NAME")
    input_start = request.args.get("start")
    input_end = request.args.get("end")

    try:
        start_date = parse_request_datetime(f"{input_start} 00:00:00")
        end_date = parse_request_datetime(f"{input_end} 23:59:59")
    except Exception as e:
        return jsonify({"error": f"Invalid date format: {e}"}), 400

    patient_table = dynamodb.Table(patient_table_name)
    appointment_table = dynamodb.Table(appointment_table_name)

    patients = scan_table_by_created_at(patient_table, start_date, end_date)
    appointments = scan_table_by_created_at(appointment_table, start_date, end_date)

    df_patients = pd.DataFrame(patients)
    df_appointments = pd.DataFrame(appointments)
    if df_patients.empty or df_appointments.empty:
        return jsonify([])

    df_patients['old_appointment_id'] = df_patients['old_appointment_id'].astype(str)
    df_appointments['id'] = df_appointments['id'].astype(str)

    merged_df = pd.merge(df_appointments, df_patients, left_on='id', right_on='old_appointment_id', how='left')
    merged_df = merged_df.fillna('')
    merged_df['Status'] = merged_df['reach_out_count'].map({
        -1: "Booked",
        -2: "Max Reached Out",
        -4: "No Selection Made",
        -3: "No Available Slots",
        -5: "Failed",
         1: "Reached Once",
         2: "Reached Twice"
    }).fillna("Other")
    merged_df['Provider Name'] = merged_df['doctor_name'].fillna('').astype(str).str.strip().replace('', 'Unknown')

    booked = merged_df[merged_df['Status'] == 'Booked']
    # Deduplicate using fields that are always present in merged_df
    booked = booked.drop_duplicates(subset=["doctor_name", "reach_out_count", "created_at_x"])
    if booked.empty:
        return jsonify([])

    counts = booked.groupby('Provider Name').size().reset_index(name='Booked')
    return jsonify(counts.to_dict(orient='records'))


def get_db_connection():
    return psycopg2.connect(os.getenv("DATABASE_URL"), cursor_factory=RealDictCursor)

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    email = data.get('email')
    password = data.get('password')

    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400

    allowed_domains = {"ohcc.org", "naveonguides.com"}
    domain = email.lower().split('@')[-1]
    if domain not in allowed_domains:
        return jsonify({"error": "Registration is restricted to authorized email domains."}), 403

    tenant_id = email.split('@')[1].split('.')[0]
    password_hash = generate_password_hash(password)

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO users (email, password_hash, tenant_id) VALUES (%s, %s, %s)",
                    (email, password_hash, tenant_id))
        conn.commit()
        cur.close()
        conn.close()
    except psycopg2.errors.UniqueViolation:
        return jsonify({"error": "User already exists"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"message": "User registered successfully", "tenant_id": tenant_id})

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email')
    password = data.get('password')

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if not user or not check_password_hash(user['password_hash'], password):
            return jsonify({"error": "Invalid credentials"}), 401

        return jsonify({"message": "Login successful", "tenant_id": user['tenant_id']})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/ai-show-rate')
def ai_show_rate():
    """
    Returns booked / kept counts (and % show-rate) for one or more months.
    Skips any appointment dated today or in the future unless its status is already 'Kept'.
    """
    months = request.args.getlist("months")
    if not months:
        return jsonify({"error": "At least one month must be specified as 'YYYY-MM'."}), 400

    try:
        for month_str in months:
            month_start_end(month_str)
    except ValueError:
        return jsonify({"error": "Month must be in YYYY-MM format."}), 400

    return jsonify(calculate_ai_show_rate(months))


@app.route('/reports/monthly-metrics.xlsx')
def monthly_metrics_report():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    start_month = request.args.get("start", "2025-01")
    end_month = request.args.get("end", last_completed_month())

    try:
        months = list(iter_months(start_month, end_month))
    except Exception:
        return jsonify({"error": "start and end must be in YYYY-MM format."}), 400

    if not months:
        return jsonify({"error": "Report range must include at least one month."}), 400
    if len(months) > 36:
        return jsonify({"error": "Report range cannot exceed 36 months."}), 400

    ai_items = scan_ai_booking_items()

    rows = []
    for month_str in months:
        month_start, month_end = month_start_end(month_str)
        month_records = build_live_detail_records(month_start, month_end)
        dashboard = calculate_dashboard_metrics(month_records)
        ai = calculate_ai_show_rate([month_str], ai_items)
        ai_rate = ai["show_rate"] / 100
        rows.append({
            "year": int(month_str[:4]),
            "month": month_str,
            "label": datetime.strptime(month_str, "%Y-%m").strftime("%b %Y"),
            "total_sent": dashboard["total_sent"],
            "total_booked": dashboard["total_booked"],
            "percent_booked": dashboard["percent_booked"],
            "estimated_revenue": dashboard["estimated_revenue"],
            "ai_show_rate": ai_rate,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Metrics"
    summary = wb.create_sheet("Summary")
    notes = wb.create_sheet("Source Notes")
    ws.sheet_view.showGridLines = False
    summary.sheet_view.showGridLines = False
    notes.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F7A83"
    summary.sheet_properties.tabColor = "145C63"
    notes.sheet_properties.tabColor = "7DB7B1"

    primary = "145C63"
    header_fill = "D8ECE9"
    total_fill = "EEF6F5"
    note_fill = "E7F3F1"
    border_color = "C9D7D5"
    body_fill = "FBFEFE"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    headers = [
        "Year",
        "Month",
        "Total Sent",
        "Total Booked",
        "% Booked",
        "Estimated Revenue",
        "AI Show Rate %",
        "Adjusted Revenue",
    ]

    ws.merge_cells("A1:H1")
    ws["A1"] = "BookIt Dashboard Monthly Metrics"
    ws["A1"].fill = PatternFill("solid", fgColor=primary)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated on demand | Range: {start_month} through {end_month}"
    ws["A2"].fill = PatternFill("solid", fgColor=note_fill)
    ws["A2"].font = Font(color="46666A", italic=True, size=10)

    ws.append([])
    ws.append(headers)
    header_row = 4
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=True, color="173234")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 26

    data_start = 5
    for item in rows:
        ws.append([
            item["year"],
            item["label"],
            item["total_sent"],
            item["total_booked"],
            None,
            item["estimated_revenue"],
            item["ai_show_rate"],
            None,
        ])
        current_row = ws.max_row
        ws[f"E{current_row}"] = f"=IFERROR(D{current_row}/C{current_row},0)"
        ws[f"H{current_row}"] = f"=F{current_row}*G{current_row}"

    data_end = ws.max_row
    total_rows = {}
    for year in sorted({item["year"] for item in rows}):
        ws.append([f"{year} Total", "", None, None, None, None, None, None])
        row = ws.max_row
        total_rows[year] = row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f"C{row}"] = f'=SUMIF(A{data_start}:A{data_end},{year},C{data_start}:C{data_end})'
        ws[f"D{row}"] = f'=SUMIF(A{data_start}:A{data_end},{year},D{data_start}:D{data_end})'
        ws[f"E{row}"] = f"=IFERROR(D{row}/C{row},0)"
        ws[f"F{row}"] = f'=SUMIF(A{data_start}:A{data_end},{year},F{data_start}:F{data_end})'
        ws[f"G{row}"] = f"=IFERROR(H{row}/F{row},0)"
        ws[f"H{row}"] = f'=SUMIF(A{data_start}:A{data_end},{year},H{data_start}:H{data_end})'

    ws.append(["Grand Total", "", None, None, None, None, None, None])
    grand_total_row = ws.max_row
    ws.merge_cells(start_row=grand_total_row, start_column=1, end_row=grand_total_row, end_column=2)
    first_total_row = min(total_rows.values())
    last_total_row = max(total_rows.values())
    ws[f"C{grand_total_row}"] = f"=SUM(C{first_total_row}:C{last_total_row})"
    ws[f"D{grand_total_row}"] = f"=SUM(D{first_total_row}:D{last_total_row})"
    ws[f"E{grand_total_row}"] = f"=IFERROR(D{grand_total_row}/C{grand_total_row},0)"
    ws[f"F{grand_total_row}"] = f"=SUM(F{first_total_row}:F{last_total_row})"
    ws[f"G{grand_total_row}"] = f"=IFERROR(H{grand_total_row}/F{grand_total_row},0)"
    ws[f"H{grand_total_row}"] = f"=SUM(H{first_total_row}:H{last_total_row})"

    for row in ws.iter_rows(min_row=header_row, max_row=grand_total_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for row in range(data_start, data_end + 1):
        fill = PatternFill("solid", fgColor=body_fill if row % 2 else "F6FBFA")
        for cell in ws[row]:
            cell.fill = fill

    for row in range(data_start, grand_total_row + 1):
        ws[f"C{row}"].number_format = '#,##0'
        ws[f"D{row}"].number_format = '#,##0'
        ws[f"E{row}"].number_format = '0.0%'
        ws[f"F{row}"].number_format = '$#,##0'
        ws[f"G{row}"].number_format = '0.0%'
        ws[f"H{row}"].number_format = '$#,##0'

    for row in range(data_end + 1, grand_total_row + 1):
        for cell in ws[row]:
            cell.fill = PatternFill("solid", fgColor=total_fill)
            cell.font = Font(bold=True, color="173234")

    for col, width in {
        "A": 12,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 18,
        "G": 16,
        "H": 18,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:H{grand_total_row}"
    if data_start <= data_end:
        ws.conditional_formatting.add(
            f"E{data_start}:E{data_end}",
            DataBarRule(start_type="min", end_type="max", color="7DB7B1", showValue=True),
        )
        ws.conditional_formatting.add(
            f"H{data_start}:H{data_end}",
            ColorScaleRule(
                start_type="min",
                start_color="FDE68A",
                mid_type="percentile",
                mid_value=50,
                mid_color="D8ECE9",
                end_type="max",
                end_color="4F9D95",
            ),
        )

    summary.merge_cells("A1:H1")
    summary["A1"] = "BookIt Monthly Performance Summary"
    summary["A1"].fill = PatternFill("solid", fgColor=primary)
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.row_dimensions[1].height = 28
    summary.merge_cells("A2:H2")
    summary["A2"] = f"Dashboard data pulled from deployed backend. Generated workbook from {start_month} through {end_month}."
    summary["A2"].fill = PatternFill("solid", fgColor=note_fill)
    summary["A2"].font = Font(color="46666A", italic=True, size=10)
    summary.append([])
    summary.append(["Period", "Total Sent", "Total Booked", "Adjusted Revenue"])
    summary_header_row = 4
    for cell in summary[summary_header_row]:
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=True)
        cell.border = thin_border
    for year, row in total_rows.items():
        summary.append([str(year), f"='Monthly Metrics'!C{row}", f"='Monthly Metrics'!D{row}", f"='Monthly Metrics'!H{row}"])
    summary.append(["All Months", f"='Monthly Metrics'!C{grand_total_row}", f"='Monthly Metrics'!D{grand_total_row}", f"='Monthly Metrics'!H{grand_total_row}"])
    for row in summary.iter_rows(min_row=summary_header_row, max_row=summary.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    for row in range(summary_header_row + 1, summary.max_row + 1):
        summary[f"B{row}"].number_format = '#,##0'
        summary[f"C{row}"].number_format = '#,##0'
        summary[f"D{row}"].number_format = '$#,##0'
    for col, width in {"A": 18, "B": 14, "C": 14, "D": 18}.items():
        summary.column_dimensions[col].width = width

    line = LineChart()
    line.title = "Adjusted Revenue by Month"
    line.y_axis.title = "Adjusted Revenue"
    line.x_axis.title = "Month"
    data = Reference(ws, min_col=8, min_row=header_row, max_row=data_end)
    cats = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height = 8
    line.width = 20
    summary.add_chart(line, "A9")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Total Booked by Month"
    bar.y_axis.title = "Booked"
    data = Reference(ws, min_col=4, min_row=header_row, max_row=data_end)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 20
    summary.add_chart(bar, "A25")

    notes.merge_cells("A1:B1")
    notes["A1"] = "Source Notes and Calculation Definitions"
    notes["A1"].fill = PatternFill("solid", fgColor=primary)
    notes["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    notes["A1"].alignment = Alignment(horizontal="center")
    note_rows = [
        ("Backend URL", request.host_url.rstrip("/")),
        ("Data Range", f"{start_month} through {end_month}"),
        ("Dashboard Endpoint Logic", "/live-details calendar-month records with dashboard deduplication."),
        ("AI Show Rate Logic", "/ai-show-rate by calendar month, percentage only."),
        ("Total Booked", "Booked rows deduplicated by Patient ID + Date + Time + Status."),
        ("Estimated Revenue", "Booked before 2025-07-01 at $154; booked on/after 2025-07-01 at $175."),
        ("Adjusted Revenue", "Estimated Revenue x AI Show Rate %."),
        ("Generated", datetime.now(ZoneInfo(os.getenv("CLINIC_TIMEZONE", "America/Chicago"))).strftime("%Y-%m-%d %H:%M:%S %Z")),
    ]
    for row in note_rows:
        notes.append(row)
    for row in notes.iter_rows(min_row=2, max_row=notes.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].fill = PatternFill("solid", fgColor=header_fill)
        row[0].font = Font(bold=True)
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"bookit_monthly_metrics_{start_month}_to_{end_month}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# --- AWS Credentials Debug Route ---
@app.route('/debug-aws')
def debug_aws():
    import boto3
    from botocore.exceptions import ClientError

    try:
        sts = boto3.client('sts')
        identity = sts.get_caller_identity()
        return jsonify({
            "Account": identity.get("Account"),
            "UserId": identity.get("UserId"),
            "Arn": identity.get("Arn")
        })
    except ClientError as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
